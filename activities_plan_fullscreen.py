"""
Activity Plan Fullscreen Editor
- Fullscreen Tkinter dashboard that loads "Activity Plan.xlsx" (same folder by default)
- Visualizes merged cells, background fills, bold headers
- Inline editing (double-click), auto-save after edit
- Add Row (append), Mark as Done, Week filter (show only rows in chosen week)
- Shortcuts: Ctrl+S save, F11 toggle fullscreen, Esc cancel edit

Dependencies:
    pip install openpyxl pillow python-dateutil
"""
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color
from datetime import datetime, timedelta, date
from dateutil.parser import parse as dtparse

# --------------- Config ---------------
DEFAULT_FILENAME = "Activity Plan.xlsx"
CELL_MIN_WIDTH = 10
CELL_MIN_HEIGHT = 1
# Expected column header names (case-insensitive) for special features
DATE_COL_NAMES = ["date", "task date", "start date"]
STATUS_COL_NAMES = ["status"]
# --------------------------------------


def hex_from_openpyxl_color(color):
    if color is None:
        return None
    # Try to get the RGB string safely
    rgb = color.rgb if hasattr(color, "rgb") else None
    if rgb is None:
        return None
    if isinstance(rgb, bytes):
        rgb = rgb.decode()

    # If it's ARGB (8 chars), strip the alpha
    if isinstance(rgb, str) and len(rgb) == 8:
        rgb = rgb[2:]
    return f"#{rgb}" if rgb else None



class ActivityApp(tk.Tk):
    def __init__(self, path=None):
        super().__init__()
        self.title("Activity Plan — Fullscreen Editor")
        self.fullscreen = True
        self.attributes("-fullscreen", True)

        # file
        self.path = path if path and os.path.exists(path) else (DEFAULT_FILENAME if os.path.exists(DEFAULT_FILENAME) else None)
        if not self.path:
            # ask user to choose file
            self.path = filedialog.askopenfilename(title="Open Activity Plan Excel", filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx")])
            if not self.path:
                messagebox.showerror("No file", "No Excel file selected. Exiting.")
                self.destroy()
                return

        # workbook and sheet
        self.wb = load_workbook(self.path)
        self.sheet = self.wb.active

        # mapping structures
        self.merged_map = {}       # (min_row, min_col) -> (max_row, max_col)
        self.cell_widgets = {}     # (r,c) -> Label widget
        self.visible_cells = set() # currently visible widget coords (for filtering)
        self.edit_overlay = None   # Entry overlay for editing
        self.edit_target = None    # (r,c)

        # column indices
        self.header_map = {}  # normalized header -> col index
        self.date_col = None
        self.status_col = None

        # UI
        self._build_ui()
        self._bind_shortcuts()
        self._scan_sheet()
        self._render_sheet()

    # ---------------- UI ----------------
    def _build_ui(self):
        # toolbar top
        toolbar = ttk.Frame(self)
        toolbar.pack(side="top", fill="x", padx=4, pady=4)

        ttk.Button(toolbar, text="Open...", command=self.open_file).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Save (Ctrl+S)", command=self.save).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Add Row", command=self.add_row).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Mark Selected Done", command=self.mark_selected_done).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Exit Fullscreen (F11)", command=self.toggle_fullscreen).pack(side="left", padx=4)

        # week filter inputs
        ttk.Label(toolbar, text="Week start (YYYY-MM-DD):").pack(side="left", padx=(18,4))
        self.week_entry = ttk.Entry(toolbar, width=12)
        self.week_entry.pack(side="left")
        ttk.Button(toolbar, text="Filter Week", command=self.filter_week).pack(side="left", padx=4)
        ttk.Button(toolbar, text="Clear Filter", command=self.clear_filter).pack(side="left", padx=4)

        # status label
        self.status_var = tk.StringVar(value=f"Loaded: {os.path.basename(self.path)}")
        ttk.Label(toolbar, textvariable=self.status_var).pack(side="right", padx=8)

        # container with canvas and scrollbars
        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(container, bg="#f6f6f6")
        self.vbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.hbar = ttk.Scrollbar(container, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vbar.set, xscrollcommand=self.hbar.set)
        self.vbar.pack(side="right", fill="y")
        self.hbar.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.grid_frame = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.grid_frame, anchor="nw")
        self.grid_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        # a Tree-like selection: we'll use a simple selected cell state
        self.selected_cell = None

    def _bind_shortcuts(self):
        self.bind("<Control-s>", lambda e: self.save())
        self.bind("<Escape>", lambda e: self._cancel_edit())
        self.bind("<F11>", lambda e: self.toggle_fullscreen())

    # ------------- Workbook helpers -------------
    def _scan_sheet(self):
        # build merged map
        self.merged_map.clear()
        for mg in self.sheet.merged_cells.ranges:
            self.merged_map[(mg.min_row, mg.min_col)] = (mg.max_row, mg.max_col)

        # header map: read first ~10 rows to find header names and their columns
        self.header_map.clear()
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        # Heuristic: headers are within first 8 rows; search for cells with text that matches "Date" or "Status"
        for r in range(1, min(10, max_row) + 1):
            for c in range(1, max_col + 1):
                val = self.sheet.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    key = val.strip().lower()
                    if key not in self.header_map:
                        self.header_map[key] = c

        # find date and status columns by common names
        self.date_col = None
        self.status_col = None
        for name in DATE_COL_NAMES:
            if name in self.header_map:
                self.date_col = self.header_map[name]
                break
        for name in STATUS_COL_NAMES:
            if name in self.header_map:
                self.status_col = self.header_map[name]
                break

    def _find_merged_span_top_left(self, r, c):
        # return ((min_row,min_col),(max_row,max_col)) if (r,c) is the top-left of a merged block
        for (mr, mc), (Mr, Mc) in self.merged_map.items():
            if (r, c) == (mr, mc):
                return ((mr, mc), (Mr, Mc))
            if mr <= r <= Mr and mc <= c <= Mc:
                # inner cell
                return ((mr, mc), (Mr, Mc))
        return None

    # ------------- Rendering -------------
    def _render_sheet(self):
        # clear grid_frame
        for w in self.grid_frame.winfo_children():
            w.destroy()
        self.cell_widgets.clear()
        self.visible_cells.clear()

        max_row = self.sheet.max_row
        max_col = self.sheet.max_column

        # place labels respecting merged cells
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                # skip if part of a merged region that isn't the top-left
                span = self._find_merged_span_top_left(r, c)
                if span:
                    (mr, mc), (Mr, Mc) = span
                    if (r, c) != (mr, mc):
                        continue
                    rowspan = Mr - mr + 1
                    colspan = Mc - mc + 1
                else:
                    rowspan = 1
                    colspan = 1

                cell = self.sheet.cell(row=r, column=c)
                display = "" if cell.value is None else str(cell.value)
                bg = hex_from_openpyxl_color(cell.fill.start_color) if cell.fill else None
                bold = bool(cell.font and cell.font.bold)

                lbl = tk.Label(self.grid_frame, text=display, borderwidth=1, relief="solid", anchor="w")
                # style label
                font_family = "TkDefaultFont"
                font_size = 10
                if bold:
                    lbl.config(font=(font_family, font_size, "bold"))
                else:
                    lbl.config(font=(font_family, font_size))
                if bg:
                    try:
                        lbl.config(bg=bg)
                    except Exception:
                        pass
                lbl.grid(row=r, column=c, rowspan=rowspan, columnspan=colspan, sticky="nsew", padx=0, pady=0)
                lbl._coord = (r, c)
                lbl._merged_span = ((mr, mc), (Mr, Mc)) if span else None
                # mark editable heuristic: rows after header region and non-bold
                editable = not bold and r > self._guess_header_end_row()
                lbl.editable = editable
                if editable:
                    lbl.bind("<Double-Button-1>", self._on_cell_double_click)
                    lbl.bind("<Button-1>", self._on_cell_click)
                    lbl.config(cursor="xterm")
                else:
                    lbl.config(cursor="arrow")
                # visual selection
                lbl.bind("<Button-3>", lambda e, w=lbl: self._show_context_menu(w, e))

                # keep references
                self.cell_widgets[(r, c)] = lbl
                self.visible_cells.add((r, c))

                # make grid resizeable
                self.grid_frame.grid_rowconfigure(r, weight=1)
                self.grid_frame.grid_columnconfigure(c, weight=1)

        # update scroll region
        self.grid_frame.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        self.set_status(f"Rendered {os.path.basename(self.path)} — {self.sheet.title}")

    def _guess_header_end_row(self):
        """
        Heuristic to find last header row: find first row that contains non-bold text in many cells.
        """
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        for r in range(1, min(12, max_row) + 1):
            non_bold_count = 0
            total = 0
            for c in range(1, max_col + 1):
                total += 1
                cell = self.sheet.cell(row=r, column=c)
                if not (cell.font and cell.font.bold):
                    non_bold_count += 1
            # if the row has many non-bold cells, treat it as data row
            if non_bold_count / total > 0.5:
                return max(1, r - 1)  # header ends at previous row
        return 5

    # ------------- Interaction -------------
    def _on_cell_click(self, event):
        widget = event.widget
        self._select_widget(widget)

    def _on_cell_double_click(self, event):
        widget = event.widget
        if not getattr(widget, "editable", False):
            return
        self._select_widget(widget)
        self._start_edit(widget)

    def _select_widget(self, widget):
        # clear previous selection
        if self.selected_cell and self.selected_cell.winfo_exists():
            self.selected_cell.config(relief="solid")
        self.selected_cell = widget
        widget.config(relief="ridge", bd=2)

    def _start_edit(self, widget):
        # overlay an Entry on top of the label
        if self.edit_overlay:
            self._cancel_edit()
        r, c = widget._coord
        x = widget.winfo_rootx() - self.winfo_rootx()
        y = widget.winfo_rooty() - self.winfo_rooty()
        w = widget.winfo_width()
        h = widget.winfo_height()

        # create overlay entry
        self.edit_overlay = tk.Entry(self.grid_frame)
        self.edit_overlay.insert(0, widget.cget("text"))
        self.edit_overlay.place(x=widget.winfo_x(), y=widget.winfo_y(), width=w, height=h)
        self.edit_overlay.focus_set()
        self.edit_target = (r, c)
        # commit on Enter
        self.edit_overlay.bind("<Return>", lambda e: self._commit_edit())
        # cancel on Esc
        self.edit_overlay.bind("<Escape>", lambda e: self._cancel_edit())

    def _cancel_edit(self):
        if self.edit_overlay:
            try:
                self.edit_overlay.destroy()
            except Exception:
                pass
            self.edit_overlay = None
            self.edit_target = None

    def _commit_edit(self):
        if not self.edit_overlay or not self.edit_target:
            return
        newval = self.edit_overlay.get()
        r, c = self.edit_target
        # set displayed text
        lbl = self.cell_widgets.get((r, c))
        if lbl:
            lbl.config(text=newval)
        # write to workbook cell (preserve type where possible)
        cell = self.sheet.cell(row=r, column=c)
        # attempt numeric conversion
        if newval.strip() == "":
            cell.value = None
        else:
            # try integer then float then date then string
            try:
                if newval.isdigit():
                    cell.value = int(newval)
                else:
                    # try float
                    f = float(newval)
                    cell.value = f
            except Exception:
                # try date parse
                try:
                    dt = dtparse(newval)
                    cell.value = dt
                except Exception:
                    cell.value = newval
        # auto-save on edit
        try:
            self.wb.save(self.path)
            self.set_status(f"Auto-saved after editing {get_column_letter(c)}{r}")
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save: {e}")
        # remove overlay
        self._cancel_edit()

    # ------------- Actions -------------
    def save(self):
        try:
            self.wb.save(self.path)
            self.set_status(f"Saved {os.path.basename(self.path)}")
            messagebox.showinfo("Saved", f"Saved {self.path}")
        except Exception as e:
            messagebox.showerror("Save error", str(e))

    def open_file(self):
        f = filedialog.askopenfilename(title="Open Activity Plan Excel", filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx")])
        if not f:
            return
        self.path = f
        try:
            self.wb = load_workbook(self.path)
            self.sheet = self.wb.active
            self._scan_sheet()
            self._render_sheet()
            self.set_status(f"Opened {os.path.basename(self.path)}")
        except Exception as e:
            messagebox.showerror("Open failed", str(e))

    def refresh(self):
        # reload file from disk then re-render
        try:
            self.wb = load_workbook(self.path)
            self.sheet = self.wb.active
            self._scan_sheet()
            self._render_sheet()
            self.set_status("Refreshed view")
        except Exception as e:
            messagebox.showerror("Refresh failed", str(e))

    def add_row(self):
        # append a blank row after last row
        max_row = self.sheet.max_row
        new_row = max_row + 1
        # create empty cells for all columns
        for c in range(1, self.sheet.max_column + 1):
            self.sheet.cell(row=new_row, column=c).value = ""
        # save and refresh
        self.wb.save(self.path)
        self.refresh()
        # optionally scroll to new row and start edit on first editable column
        # find first non-bold column in header area to edit
        self.set_status(f"Added row {new_row} and saved")

    def _show_context_menu(self, widget, event):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Edit", command=lambda w=widget: (self._select_widget(w), self._start_edit(w)))
        menu.add_command(label="Mark Done", command=lambda w=widget: (self._select_widget(w), self._mark_widget_row_done(w)))
        menu.add_command(label="Clear Cell", command=lambda w=widget: (self._select_widget(w), self._clear_widget_cell(w)))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _clear_widget_cell(self, widget):
        r, c = widget._coord
        widget.config(text="")
        self.sheet.cell(row=r, column=c).value = None
        self.wb.save(self.path)
        self.set_status(f"Cleared {get_column_letter(c)}{r}")

    def _mark_widget_row_done(self, widget):
        # find status column
        if not self.status_col:
            # try to find header named 'Status' now
            for key, col in self.header_map.items():
                if "status" in key:
                    self.status_col = col
                    break
        if not self.status_col:
            messagebox.showwarning("Missing Status column", "Could not find a Status column to mark as done.")
            return
        r, _ = widget._coord
        # set value in sheet
        self.sheet.cell(row=r, column=self.status_col).value = "Done"
        # update label in UI (if exists)
        status_lbl = self.cell_widgets.get((r, self.status_col))
        if status_lbl:
            status_lbl.config(text="Done", bg="#d0ffd0")
        try:
            self.wb.save(self.path)
            self.set_status(f"Marked row {r} as Done and saved")
        except Exception as e:
            messagebox.showerror("Save error", str(e))

    def mark_selected_done(self):
        if not self.selected_cell:
            messagebox.showinfo("Select row", "Click a cell in the row you want to mark as done.")
            return
        self._mark_widget_row_done(self.selected_cell)

    # ------------- Filtering -------------
    def filter_week(self):
        text = self.week_entry.get().strip()
        if not text:
            messagebox.showinfo("Week filter", "Enter a week start date (YYYY-MM-DD).")
            return
        try:
            start = dtparse(text).date()
        except Exception:
            messagebox.showerror("Invalid date", "Could not parse the date. Use YYYY-MM-DD or similar.")
            return
        end = start + timedelta(days=6)
        # which column is Date?
        if not self.date_col:
            # look for any header containing 'date'
            for key, col in self.header_map.items():
                if 'date' in key:
                    self.date_col = col
                    break
        if not self.date_col:
            messagebox.showwarning("Missing Date column", "Could not find a Date column for filtering.")
            return

        # show/hide labels based on date column values
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        # Hide all first
        for (r, c), widget in list(self.cell_widgets.items()):
            widget.grid_remove()
            self.visible_cells.discard((r, c))

        # iterate rows, check date cell value
        for r in range(1, max_row + 1):
            raw = self.sheet.cell(row=r, column=self.date_col).value
            if raw is None:
                display_date = None
            else:
                try:
                    if isinstance(raw, (datetime, date)):
                        display_date = raw.date() if isinstance(raw, datetime) else raw
                    else:
                        display_date = dtparse(str(raw)).date()
                except Exception:
                    display_date = None
            if display_date and start <= display_date <= end:
                # unhide entire row (all columns that have widgets)
                for c in range(1, max_col + 1):
                    w = self.cell_widgets.get((r, c))
                    if w:
                        w.grid()
                        self.visible_cells.add((r, c))
        self.set_status(f"Filtered week {start.isoformat()} to {end.isoformat()}")

    def clear_filter(self):
        # unhide every widget
        for (r, c), w in list(self.cell_widgets.items()):
            w.grid()
            self.visible_cells.add((r, c))
        self.set_status("Cleared filter")

    # ------------- Utility -------------
    def toggle_fullscreen(self):
        self.fullscreen = not self.fullscreen
        self.attributes("-fullscreen", self.fullscreen)
        if self.fullscreen:
            self.set_status("Fullscreen mode")
        else:
            self.set_status("Windowed mode")

    def set_status(self, text):
        self.status_var.set(text)


def main():
    path = DEFAULT_FILENAME if os.path.exists(DEFAULT_FILENAME) else None
    app = ActivityApp(path=path)
    # if the app failed to start (no file), ActivityApp will destroy itself; check
    try:
        app.mainloop()
    except tk.TclError:
        pass


if __name__ == "__main__":
    main()
