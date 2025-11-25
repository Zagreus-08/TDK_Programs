"""
Professional Gantt Chart Generator - GUI Version
Easy-to-use interface for managing projects and generating charts
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import datetime, timedelta
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = 'gantt_projects.json'

class GanttChartGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 Gantt Chart Generator Pro")
        self.root.geometry("1100x700")
        self.root.configure(bg='#f5f5f5')
        
        # Load data
        self.data = self.load_projects()
        
        # Style configuration
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.configure_styles()
        
        # Create UI
        self.create_ui()
        self.refresh_project_list()
        
    def configure_styles(self):
        self.style.configure('Title.TLabel', font=('Segoe UI', 18, 'bold'), background='#1565C0', foreground='white')
        self.style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'), background='#f5f5f5')
        self.style.configure('TButton', font=('Segoe UI', 10), padding=8)
        self.style.configure('Action.TButton', font=('Segoe UI', 10, 'bold'), padding=10)
        self.style.configure('Treeview', font=('Segoe UI', 9), rowheight=28)
        self.style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'))
        
    def create_ui(self):
        # Title bar
        title_frame = tk.Frame(self.root, bg='#1565C0', height=60)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="📊 Gantt Chart Generator Pro", 
                              font=('Segoe UI', 18, 'bold'), bg='#1565C0', fg='white')
        title_label.pack(side='left', padx=20, pady=15)
        
        # Main container
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # Left panel - Projects
        left_frame = ttk.LabelFrame(main_frame, text="Projects", padding=10)
        left_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        # Project buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Button(btn_frame, text="➕ Add Project", command=self.add_project).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="✏️ Edit", command=self.edit_project).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="🗑️ Delete", command=self.delete_project).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="👁️ Toggle Visible", command=self.toggle_visibility).pack(side='left', padx=2)
        
        # Project list
        self.project_tree = ttk.Treeview(left_frame, columns=('name', 'tasks', 'visible'), show='headings', height=8)
        self.project_tree.heading('name', text='Project Name')
        self.project_tree.heading('tasks', text='Tasks')
        self.project_tree.heading('visible', text='Visible')
        self.project_tree.column('name', width=250)
        self.project_tree.column('tasks', width=60, anchor='center')
        self.project_tree.column('visible', width=70, anchor='center')
        self.project_tree.pack(fill='both', expand=True)
        self.project_tree.bind('<<TreeviewSelect>>', self.on_project_select)
        
        # Scrollbar for projects
        proj_scroll = ttk.Scrollbar(left_frame, orient='vertical', command=self.project_tree.yview)
        self.project_tree.configure(yscrollcommand=proj_scroll.set)

        # Right panel - Tasks
        right_frame = ttk.LabelFrame(main_frame, text="Tasks", padding=10)
        right_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))
        
        # Task buttons
        task_btn_frame = ttk.Frame(right_frame)
        task_btn_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Button(task_btn_frame, text="➕ Add Task", command=self.add_task).pack(side='left', padx=2)
        ttk.Button(task_btn_frame, text="✏️ Edit Task", command=self.edit_task).pack(side='left', padx=2)
        ttk.Button(task_btn_frame, text="🗑️ Delete Task", command=self.delete_task).pack(side='left', padx=2)
        ttk.Button(task_btn_frame, text="⬆️ Move Up", command=self.move_task_up).pack(side='left', padx=2)
        ttk.Button(task_btn_frame, text="⬇️ Move Down", command=self.move_task_down).pack(side='left', padx=2)
        
        # Task list
        self.task_tree = ttk.Treeview(right_frame, columns=('id', 'desc', 'start', 'end', 'status'), show='headings', height=8)
        self.task_tree.heading('id', text='ID')
        self.task_tree.heading('desc', text='Description')
        self.task_tree.heading('start', text='Start')
        self.task_tree.heading('end', text='End')
        self.task_tree.heading('status', text='Status')
        self.task_tree.column('id', width=60)
        self.task_tree.column('desc', width=280)
        self.task_tree.column('start', width=70, anchor='center')
        self.task_tree.column('end', width=70, anchor='center')
        self.task_tree.column('status', width=80, anchor='center')
        self.task_tree.pack(fill='both', expand=True)
        
        # Scrollbar for tasks
        task_scroll = ttk.Scrollbar(right_frame, orient='vertical', command=self.task_tree.yview)
        self.task_tree.configure(yscrollcommand=task_scroll.set)
        
        # Bottom panel - Settings & Generate
        bottom_frame = ttk.Frame(self.root, padding=10)
        bottom_frame.pack(fill='x')
        
        # Settings
        settings_frame = ttk.LabelFrame(bottom_frame, text="Chart Settings", padding=10)
        settings_frame.pack(side='left', fill='x', expand=True)
        
        ttk.Label(settings_frame, text="Start Date:").pack(side='left', padx=5)
        self.start_date_var = tk.StringVar(value=self.data['settings']['start_date'])
        self.start_date_entry = ttk.Entry(settings_frame, textvariable=self.start_date_var, width=12)
        self.start_date_entry.pack(side='left', padx=5)
        
        ttk.Label(settings_frame, text="Weeks:").pack(side='left', padx=5)
        self.weeks_var = tk.StringVar(value=str(self.data['settings']['num_weeks']))
        self.weeks_spin = ttk.Spinbox(settings_frame, from_=4, to=12, textvariable=self.weeks_var, width=5)
        self.weeks_spin.pack(side='left', padx=5)
        
        ttk.Label(settings_frame, text="WW1 Date:").pack(side='left', padx=5)
        self.ww1_date_var = tk.StringVar(value=self.data['settings'].get('ww1_date', '2025-04-07'))
        self.ww1_date_entry = ttk.Entry(settings_frame, textvariable=self.ww1_date_var, width=12)
        self.ww1_date_entry.pack(side='left', padx=5)
        
        ttk.Button(settings_frame, text="💾 Save Settings", command=self.save_settings).pack(side='left', padx=10)
        
        # Generate buttons
        generate_frame = ttk.Frame(bottom_frame)
        generate_frame.pack(side='right', padx=10)
        
        # Update existing chart button
        self.update_btn = tk.Button(generate_frame, text="🔄 UPDATE CHART", 
                                    font=('Segoe UI', 11, 'bold'), bg='#2196F3', fg='white',
                                    padx=15, pady=8, command=self.update_chart)
        self.update_btn.pack(side='left', padx=5)
        
        # Generate new chart button
        self.generate_btn = tk.Button(generate_frame, text="📊 NEW CHART", 
                                      font=('Segoe UI', 11, 'bold'), bg='#4CAF50', fg='white',
                                      padx=15, pady=8, command=self.generate_chart)
        self.generate_btn.pack(side='left', padx=5)
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', padding=5)
        status_bar.pack(fill='x', side='bottom')
        
    def load_projects(self):
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {"projects": {}, "settings": {"start_date": "2025-11-10", "num_weeks": 6, "ww1_date": "2025-04-07"}}
    
    def save_projects(self):
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)
        self.status_var.set("💾 Data saved!")
        
    def refresh_project_list(self):
        for item in self.project_tree.get_children():
            self.project_tree.delete(item)
        
        for num, proj in sorted(self.data['projects'].items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
            visible = "✅ Yes" if proj['visible'] else "❌ No"
            self.project_tree.insert('', 'end', iid=num, values=(f"[{num}] {proj['name']}", len(proj['tasks']), visible))
    
    def refresh_task_list(self, proj_num):
        for item in self.task_tree.get_children():
            self.task_tree.delete(item)
        
        if proj_num and proj_num in self.data['projects']:
            for i, task in enumerate(self.data['projects'][proj_num]['tasks']):
                self.task_tree.insert('', 'end', iid=str(i), values=(
                    task['task_id'], task['description'][:50], task['start_date'], task['end_date'], task['status']
                ))
    
    def on_project_select(self, event):
        selection = self.project_tree.selection()
        if selection:
            proj_num = selection[0]
            self.refresh_task_list(proj_num)

    def add_project(self):
        dialog = ProjectDialog(self.root, "Add New Project")
        if dialog.result:
            num, name = dialog.result
            if num in self.data['projects']:
                messagebox.showerror("Error", f"Project {num} already exists!")
                return
            self.data['projects'][num] = {"name": name, "visible": True, "tasks": []}
            self.save_projects()
            self.refresh_project_list()
            self.status_var.set(f"✅ Project {num} added!")
    
    def edit_project(self):
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a project first!")
            return
        
        proj_num = selection[0]
        proj = self.data['projects'][proj_num]
        
        dialog = ProjectDialog(self.root, "Edit Project", proj_num, proj['name'])
        if dialog.result:
            new_num, new_name = dialog.result
            if new_num != proj_num:
                self.data['projects'][new_num] = self.data['projects'].pop(proj_num)
            self.data['projects'][new_num]['name'] = new_name
            self.save_projects()
            self.refresh_project_list()
            self.status_var.set(f"✅ Project updated!")
    
    def delete_project(self):
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a project first!")
            return
        
        proj_num = selection[0]
        if messagebox.askyesno("Confirm Delete", f"Delete project {proj_num} and all its tasks?"):
            del self.data['projects'][proj_num]
            self.save_projects()
            self.refresh_project_list()
            self.refresh_task_list(None)
            self.status_var.set(f"🗑️ Project {proj_num} deleted!")
    
    def toggle_visibility(self):
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a project first!")
            return
        
        proj_num = selection[0]
        self.data['projects'][proj_num]['visible'] = not self.data['projects'][proj_num]['visible']
        self.save_projects()
        self.refresh_project_list()
        status = "visible" if self.data['projects'][proj_num]['visible'] else "hidden"
        self.status_var.set(f"👁️ Project {proj_num} is now {status}")
    
    def add_task(self):
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a project first!")
            return
        
        proj_num = selection[0]
        proj = self.data['projects'][proj_num]
        next_id = f"{proj_num}.{len(proj['tasks']) + 1}"
        
        dialog = TaskDialog(self.root, "Add New Task", default_id=next_id)
        if dialog.result:
            proj['tasks'].append(dialog.result)
            self.save_projects()
            self.refresh_task_list(proj_num)
            self.refresh_project_list()
            self.status_var.set(f"✅ Task added!")
    
    def edit_task(self):
        proj_selection = self.project_tree.selection()
        task_selection = self.task_tree.selection()
        
        if not proj_selection or not task_selection:
            messagebox.showwarning("Warning", "Please select a project and task!")
            return
        
        proj_num = proj_selection[0]
        task_idx = int(task_selection[0])
        task = self.data['projects'][proj_num]['tasks'][task_idx]
        
        dialog = TaskDialog(self.root, "Edit Task", task)
        if dialog.result:
            self.data['projects'][proj_num]['tasks'][task_idx] = dialog.result
            self.save_projects()
            self.refresh_task_list(proj_num)
            self.status_var.set(f"✅ Task updated!")
    
    def delete_task(self):
        proj_selection = self.project_tree.selection()
        task_selection = self.task_tree.selection()
        
        if not proj_selection or not task_selection:
            messagebox.showwarning("Warning", "Please select a project and task!")
            return
        
        proj_num = proj_selection[0]
        task_idx = int(task_selection[0])
        
        if messagebox.askyesno("Confirm Delete", "Delete this task?"):
            del self.data['projects'][proj_num]['tasks'][task_idx]
            self.save_projects()
            self.refresh_task_list(proj_num)
            self.refresh_project_list()
            self.status_var.set(f"🗑️ Task deleted!")
    
    def move_task_up(self):
        proj_selection = self.project_tree.selection()
        task_selection = self.task_tree.selection()
        
        if not proj_selection or not task_selection:
            return
        
        proj_num = proj_selection[0]
        task_idx = int(task_selection[0])
        tasks = self.data['projects'][proj_num]['tasks']
        
        if task_idx > 0:
            tasks[task_idx], tasks[task_idx-1] = tasks[task_idx-1], tasks[task_idx]
            self.save_projects()
            self.refresh_task_list(proj_num)
            self.task_tree.selection_set(str(task_idx-1))
    
    def move_task_down(self):
        proj_selection = self.project_tree.selection()
        task_selection = self.task_tree.selection()
        
        if not proj_selection or not task_selection:
            return
        
        proj_num = proj_selection[0]
        task_idx = int(task_selection[0])
        tasks = self.data['projects'][proj_num]['tasks']
        
        if task_idx < len(tasks) - 1:
            tasks[task_idx], tasks[task_idx+1] = tasks[task_idx+1], tasks[task_idx]
            self.save_projects()
            self.refresh_task_list(proj_num)
            self.task_tree.selection_set(str(task_idx+1))
    
    def save_settings(self):
        self.data['settings']['start_date'] = self.start_date_var.get()
        self.data['settings']['num_weeks'] = int(self.weeks_var.get())
        self.data['settings']['ww1_date'] = self.ww1_date_var.get()
        self.save_projects()
        self.status_var.set("⚙️ Settings saved!")

    def generate_chart(self):
        """Create a new chart - asks user to name the file"""
        visible_projects = {k: v for k, v in self.data['projects'].items() if v['visible']}
        
        if not visible_projects:
            messagebox.showerror("Error", "No visible projects! Show at least one project.")
            return
        
        # Ask user to name the file
        filename = filedialog.asksaveasfilename(
            title="Save New Gantt Chart As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Gantt_Chart.xlsx"
        )
        
        if not filename:  # User cancelled
            return
        
        self.status_var.set("⏳ Creating new chart...")
        self.root.update()
        
        try:
            generator = GanttChartGenerator(self.data)
            generator.generate(filename)
            self.data['settings']['current_file'] = filename
            self.save_projects()
            self.status_var.set(f"✅ New chart created: {os.path.basename(filename)}")
            messagebox.showinfo("Success", f"New Gantt chart created!\n\nFile: {filename}")
            os.startfile(filename)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate chart:\n{str(e)}")
            self.status_var.set("❌ Generation failed!")
    
    def update_chart(self):
        """Update an existing chart - asks user to select the file"""
        visible_projects = {k: v for k, v in self.data['projects'].items() if v['visible']}
        
        if not visible_projects:
            messagebox.showerror("Error", "No visible projects! Show at least one project.")
            return
        
        # Get last used file as default directory
        last_file = self.data['settings'].get('current_file', '')
        initial_dir = os.path.dirname(last_file) if last_file else os.getcwd()
        
        # Ask user to select file to update
        filename = filedialog.askopenfilename(
            title="Select Gantt Chart to Update",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=initial_dir
        )
        
        if not filename:  # User cancelled
            return
        
        self.status_var.set(f"🔄 Updating {os.path.basename(filename)}...")
        self.root.update()
        
        try:
            generator = GanttChartGenerator(self.data)
            generator.generate(filename)
            self.data['settings']['current_file'] = filename
            self.save_projects()
            self.status_var.set(f"✅ Chart updated: {os.path.basename(filename)}")
            messagebox.showinfo("Success", f"Gantt chart updated!\n\nFile: {filename}")
        except PermissionError:
            messagebox.showerror("Error", f"Cannot update file!\n\nPlease close the file in Excel first, then try again.")
            self.status_var.set("❌ File is open in Excel!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update chart:\n{str(e)}")
            self.status_var.set("❌ Update failed!")


class ProjectDialog:
    def __init__(self, parent, title, num="", name=""):
        self.result = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x150")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        frame = ttk.Frame(self.dialog, padding=20)
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text="Project Number:").grid(row=0, column=0, sticky='w', pady=5)
        self.num_entry = ttk.Entry(frame, width=10)
        self.num_entry.grid(row=0, column=1, sticky='w', pady=5)
        self.num_entry.insert(0, num)
        
        ttk.Label(frame, text="Project Name:").grid(row=1, column=0, sticky='w', pady=5)
        self.name_entry = ttk.Entry(frame, width=40)
        self.name_entry.grid(row=1, column=1, sticky='w', pady=5)
        self.name_entry.insert(0, name)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="✓ Save", command=self.save).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✗ Cancel", command=self.dialog.destroy).pack(side='left', padx=5)
        
        self.dialog.wait_window()
    
    def save(self):
        num = self.num_entry.get().strip()
        name = self.name_entry.get().strip()
        
        if not num or not name:
            messagebox.showerror("Error", "Please fill in all fields!")
            return
        
        self.result = (num, name)
        self.dialog.destroy()


class TaskDialog:
    def __init__(self, parent, title, task=None, default_id=""):
        self.result = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("500x280")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        frame = ttk.Frame(self.dialog, padding=20)
        frame.pack(fill='both', expand=True)
        
        # Task ID
        ttk.Label(frame, text="Task ID:").grid(row=0, column=0, sticky='w', pady=5)
        self.id_entry = ttk.Entry(frame, width=15)
        self.id_entry.grid(row=0, column=1, sticky='w', pady=5)
        self.id_entry.insert(0, task['task_id'] if task else default_id)
        
        # Description
        ttk.Label(frame, text="Description:").grid(row=1, column=0, sticky='w', pady=5)
        self.desc_entry = ttk.Entry(frame, width=50)
        self.desc_entry.grid(row=1, column=1, sticky='w', pady=5)
        if task:
            self.desc_entry.insert(0, task['description'])
        
        # Start Date
        ttk.Label(frame, text="Start Date (DD-Mon):").grid(row=2, column=0, sticky='w', pady=5)
        self.start_entry = ttk.Entry(frame, width=15)
        self.start_entry.grid(row=2, column=1, sticky='w', pady=5)
        if task:
            self.start_entry.insert(0, task['start_date'])
        else:
            self.start_entry.insert(0, datetime.now().strftime('%d-%b'))
        
        # End Date
        ttk.Label(frame, text="End Date (DD-Mon):").grid(row=3, column=0, sticky='w', pady=5)
        self.end_entry = ttk.Entry(frame, width=15)
        self.end_entry.grid(row=3, column=1, sticky='w', pady=5)
        if task:
            self.end_entry.insert(0, task['end_date'])
        else:
            self.end_entry.insert(0, datetime.now().strftime('%d-%b'))
        
        # Status
        ttk.Label(frame, text="Status:").grid(row=4, column=0, sticky='w', pady=5)
        self.status_var = tk.StringVar(value=task['status'] if task else 'Plan')
        status_combo = ttk.Combobox(frame, textvariable=self.status_var, width=15,
                                    values=['Done', 'Ongoing', 'Plan', 'Delayed', 'Re-schedule', 'On Hold'])
        status_combo.grid(row=4, column=1, sticky='w', pady=5)
        
        # Buttons
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="✓ Save", command=self.save).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="✗ Cancel", command=self.dialog.destroy).pack(side='left', padx=5)
        
        self.dialog.wait_window()
    
    def save(self):
        task_id = self.id_entry.get().strip()
        desc = self.desc_entry.get().strip()
        start = self.start_entry.get().strip()
        end = self.end_entry.get().strip()
        status = self.status_var.get()
        
        if not all([task_id, desc, start, end]):
            messagebox.showerror("Error", "Please fill in all fields!")
            return
        
        self.result = {
            'task_id': task_id,
            'description': desc,
            'start_date': start,
            'end_date': end,
            'status': status
        }
        self.dialog.destroy()


class GanttChartGenerator:
    """Generates the Excel Gantt chart"""
    
    def __init__(self, data):
        self.data = data
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Activity Plan"
        
        self.colors = {
            'done': PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid'),
            'ongoing': PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid'),
            'plan': PatternFill(start_color='42A5F5', end_color='42A5F5', fill_type='solid'),
            'delayed': PatternFill(start_color='E53935', end_color='E53935', fill_type='solid'),
            'reschedule': PatternFill(start_color='AB47BC', end_color='AB47BC', fill_type='solid'),
            'on_hold': PatternFill(start_color='78909C', end_color='78909C', fill_type='solid'),
            'bar_done': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),
            'bar_ongoing': PatternFill(start_color='FFB74D', end_color='FFB74D', fill_type='solid'),
            'bar_plan': PatternFill(start_color='64B5F6', end_color='64B5F6', fill_type='solid'),
            'bar_delayed': PatternFill(start_color='EF5350', end_color='EF5350', fill_type='solid'),
            'header_primary': PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid'),
            'week_header': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
            'day_header': PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid'),
            'section_header': PatternFill(start_color='263238', end_color='263238', fill_type='solid'),
            'alt_row': PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid'),
            'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
            'today': PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid'),
            'weekend': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
        }
        
        self.thin_border = Border(
            left=Side(style='thin', color='BDBDBD'),
            right=Side(style='thin', color='BDBDBD'),
            top=Side(style='thin', color='BDBDBD'),
            bottom=Side(style='thin', color='BDBDBD')
        )
        
        self.current_row = 1
        self.date_col_start = 7
        self.dates = []
        
    def generate(self, filename='Professional_Gantt_Chart.xlsx'):
        self.setup_page()
        self.create_title()
        self.create_legend()
        
        start_date = datetime.strptime(self.data['settings']['start_date'], '%Y-%m-%d')
        num_weeks = self.data['settings']['num_weeks']
        
        self.create_date_headers(start_date, num_weeks)
        self.create_main_headers()
        
        visible_projects = {k: v for k, v in self.data['projects'].items() if v['visible']}
        
        for proj_num, proj in sorted(visible_projects.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
            self.add_section_header(proj_num, proj['name'])
            
            for i, task in enumerate(proj['tasks'], 1):
                try:
                    task_start = datetime.strptime(f"{task['start_date']}-2025", '%d-%b-%Y')
                    task_end = datetime.strptime(f"{task['end_date']}-2025", '%d-%b-%Y')
                except:
                    task_start = task['start_date']
                    task_end = task['end_date']
                
                self.add_activity(i, task['task_id'], task['description'], task_start, task_end, task['status'])
            
            self.add_empty_row()
        
        self.wb.save(filename)
        return filename
    
    def setup_page(self):
        self.ws.freeze_panes = 'G8'
        self.ws.sheet_view.showGridLines = False
        self.ws.column_dimensions['A'].width = 6
        self.ws.column_dimensions['B'].width = 8
        self.ws.column_dimensions['C'].width = 55
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 12
        self.ws.column_dimensions['F'].width = 12
    
    def create_title(self):
        self.ws.merge_cells('A1:F1')
        title_cell = self.ws['A1']
        title_cell.value = "ACTIVITY PLAN"
        title_cell.font = Font(name='Calibri', size=24, bold=True, color='1565C0')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[1].height = 40
        
        self.ws.merge_cells('A2:F2')
        subtitle_cell = self.ws['A2']
        subtitle_cell.value = f"Weekly Progress Tracking Dashboard | Generated: {datetime.now().strftime('%B %d, %Y')}"
        subtitle_cell.font = Font(name='Calibri', size=11, italic=True, color='757575')
        subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[2].height = 25
        self.current_row = 3
    
    def create_legend(self):
        row = self.current_row
        legend_items = [('Done', 'done', '✓'), ('Ongoing', 'ongoing', '●'), ('Planned', 'plan', '○'),
                       ('Delayed', 'delayed', '!'), ('Re-sch', 'reschedule', '↻'), ('On Hold', 'on_hold', '⏸')]
        
        col = 1
        for label, color_key, symbol in legend_items:
            cell = self.ws.cell(row=row, column=col)
            cell.value = symbol
            cell.fill = self.colors[color_key]
            cell.font = Font(size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            
            label_cell = self.ws.cell(row=row, column=col + 1)
            label_cell.value = label
            label_cell.font = Font(size=9, color='424242')
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            col += 2
        
        self.ws.row_dimensions[row].height = 22
        self.current_row = row + 2

    def create_date_headers(self, start_date, num_weeks):
        self.dates = []
        current = start_date
        for _ in range(num_weeks * 7):
            self.dates.append(current)
            current += timedelta(days=1)
        
        header_row = self.current_row
        day_row = header_row + 1
        
        col = self.date_col_start
        for week_num in range(num_weeks):
            week_start = start_date + timedelta(weeks=week_num)
            # ISO week number (Week 1 starts from first week of January)
            week_number = week_start.isocalendar()[1]
            
            start_col = get_column_letter(col)
            end_col = get_column_letter(col + 6)
            self.ws.merge_cells(f'{start_col}{header_row}:{end_col}{header_row}')
            
            week_cell = self.ws.cell(row=header_row, column=col)
            # Calculate work week based on user-defined WW1 start date
            # Default: April 7, 2025 (so Aug 11, 2025 = WW20)
            ww1_date_str = self.data['settings'].get('ww1_date', '2025-04-07')
            ww1_start = datetime.strptime(ww1_date_str, '%Y-%m-%d')
            days_diff = (week_start - ww1_start).days
            display_week = (days_diff // 7) + 1
            if display_week <= 0:
                display_week += 52  # Previous fiscal year
            elif display_week > 52:
                display_week -= 52  # Next fiscal year
            week_cell.value = f"WW{display_week}"
            week_cell.fill = self.colors['week_header']
            week_cell.font = Font(name='Calibri', size=11, bold=True, color='1565C0')
            week_cell.alignment = Alignment(horizontal='center', vertical='center')
            week_cell.border = self.thin_border
            
            for day_offset in range(7):
                day_date = week_start + timedelta(days=day_offset)
                day_cell = self.ws.cell(row=day_row, column=col + day_offset)
                # Format: "Nov 25\nTue" - shows month, day, and weekday
                day_cell.value = day_date.strftime('%b %d\n%a')
                
                if day_date.weekday() >= 5:
                    day_cell.fill = self.colors['weekend']
                    day_cell.font = Font(size=7, color='9E9E9E')
                else:
                    day_cell.fill = self.colors['day_header']
                    day_cell.font = Font(size=7, color='424242')
                
                if day_date.date() == datetime.now().date():
                    day_cell.fill = self.colors['today']
                    day_cell.font = Font(size=7, bold=True, color='F57F17')
                
                day_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                day_cell.border = self.thin_border
                # Wider columns to fit "Nov 25"
                self.ws.column_dimensions[get_column_letter(col + day_offset)].width = 6.5
            col += 7
        
        self.ws.row_dimensions[header_row].height = 22
        self.ws.row_dimensions[day_row].height = 35
        self.current_row = day_row + 1
    
    def create_main_headers(self):
        row = self.current_row
        headers = [('No.', 'A'), ('ID', 'B'), ('Activity / Task Description', 'C'),
                   ('Start', 'D'), ('End', 'E'), ('Status', 'F')]
        
        for header_text, col_letter in headers:
            cell = self.ws[f'{col_letter}{row}']
            cell.value = header_text
            cell.fill = self.colors['header_primary']
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        self.ws.row_dimensions[row].height = 28
        self.current_row = row + 1
    
    def add_section_header(self, section_num, section_title):
        row = self.current_row
        
        num_cell = self.ws.cell(row=row, column=1)
        num_cell.value = section_num
        num_cell.fill = self.colors['section_header']
        num_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.border = self.thin_border
        
        self.ws.merge_cells(f'B{row}:F{row}')
        title_cell = self.ws.cell(row=row, column=2)
        title_cell.value = f"  {section_title}"
        title_cell.fill = self.colors['section_header']
        title_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        title_cell.border = self.thin_border
        
        for col in range(self.date_col_start, self.date_col_start + len(self.dates)):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = self.colors['section_header']
            cell.border = self.thin_border
        
        self.ws.row_dimensions[row].height = 26
        self.current_row = row + 1
    
    def add_activity(self, row_num, task_id, description, start_date, end_date, status):
        row = self.current_row
        is_alt_row = (row % 2 == 0)
        
        status_config = {
            'Done': ('done', 'bar_done', '✓', 'FFFFFF'),
            'Ongoing': ('ongoing', 'bar_ongoing', '●', 'FFFFFF'),
            'Plan': ('plan', 'bar_plan', '○', 'FFFFFF'),
            'Planned': ('plan', 'bar_plan', '○', 'FFFFFF'),
            'Delayed': ('delayed', 'bar_delayed', '!', 'FFFFFF'),
            'Re-schedule': ('reschedule', 'bar_plan', '↻', 'FFFFFF'),
            'On Hold': ('on_hold', 'bar_plan', '⏸', 'FFFFFF'),
        }
        
        status_fill, bar_color, status_symbol, text_color = status_config.get(status, ('plan', 'bar_plan', '○', 'FFFFFF'))
        row_fill = self.colors['alt_row'] if is_alt_row else self.colors['white']
        
        # Fill cells
        cells_data = [
            (1, row_num, 'center', Font(size=9, color='757575')),
            (2, task_id, 'center', Font(size=9, bold=True, color='424242')),
            (3, description, 'left', Font(size=9, color='212121')),
            (4, start_date.strftime('%d-%b') if isinstance(start_date, datetime) else start_date, 'center', Font(size=9, color='616161')),
            (5, end_date.strftime('%d-%b') if isinstance(end_date, datetime) else end_date, 'center', Font(size=9, color='616161')),
        ]
        
        for col, value, align, font in cells_data:
            cell = self.ws.cell(row=row, column=col)
            cell.value = value
            cell.fill = row_fill
            cell.font = font
            cell.alignment = Alignment(horizontal=align, vertical='center', indent=1 if col == 3 else 0)
            cell.border = self.thin_border
        
        # Status cell
        cell_status = self.ws.cell(row=row, column=6)
        cell_status.value = f"{status_symbol} {status}"
        cell_status.fill = self.colors[status_fill]
        cell_status.font = Font(size=9, bold=True, color=text_color)
        cell_status.alignment = Alignment(horizontal='center', vertical='center')
        cell_status.border = self.thin_border
        
        # Draw Gantt bar
        self.draw_gantt_bar(row, start_date, end_date, bar_color, row_fill)
        self.ws.row_dimensions[row].height = 22
        self.current_row = row + 1
    
    def draw_gantt_bar(self, row, start_date, end_date, bar_color, row_fill):
        if not self.dates:
            return
        
        if isinstance(start_date, str):
            try:
                start_date = datetime.strptime(f"{start_date}-2025", '%d-%b-%Y')
            except:
                return
        if isinstance(end_date, str):
            try:
                end_date = datetime.strptime(f"{end_date}-2025", '%d-%b-%Y')
            except:
                return
        
        for col_idx, date in enumerate(self.dates):
            cell = self.ws.cell(row=row, column=self.date_col_start + col_idx)
            if date.weekday() >= 5:
                cell.fill = self.colors['weekend']
            else:
                cell.fill = row_fill
            if date.date() == datetime.now().date():
                cell.fill = self.colors['today']
            cell.border = self.thin_border
        
        for col_idx, date in enumerate(self.dates):
            if start_date.date() <= date.date() <= end_date.date():
                cell = self.ws.cell(row=row, column=self.date_col_start + col_idx)
                cell.fill = self.colors[bar_color]
                cell.border = self.thin_border
    
    def add_empty_row(self):
        self.ws.row_dimensions[self.current_row].height = 8
        self.current_row += 1


if __name__ == "__main__":
    root = tk.Tk()
    app = GanttChartGUI(root)
    root.mainloop()
