"""
Professional Gantt Chart Generator for Excel
Creates a clean, organized Gantt chart matching the reference template
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_gantt_chart():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activity Plan"
    
    # Define colors matching the original
    colors = {
        'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'green': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'orange': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
        'light_gray': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'header_blue': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    }
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # ===== LEGEND SECTION =====
    ws['A2'] = 'Legend:'
    ws['A2'].font = Font(bold=True, size=10)
    
    legend_data = [
        ('A3', 'Plan', 'yellow'),
        ('A4', 'Re-schedule', 'yellow'),
        ('A5', 'Ongoing', 'orange'),
        ('A6', 'DONE', 'green'),
        ('A7', 'FINISHED', 'green'),
        ('A8', 'DELAYED/LEAVE', 'red'),
        ('A9', 'Editing/Delay', 'orange')
    ]
    
    for cell_ref, text, color in legend_data:
        cell = ws[cell_ref]
        cell.value = text
        cell.fill = colors[color]
        cell.font = Font(size=9, bold=False)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # ===== WEEK HEADERS (Row 10) =====
    # Starting from column F (6)
    week_start_col = 6
    weeks_data = [
        ('WW43', '21-Oct\nFri', '21-Oct\nSat', '22-Oct\nSun'),
        ('WW44', '28-Oct\nMon', '30-Oct\nTue', '04-Nov\nThu'),
        ('WW45', '04-Nov\nThu', '06-Nov\nSat', '08-Nov\nSat'),
        ('WW46', '08-Nov\nSat', '11-Nov\nTue', '13-Nov\nSat'),
    ]
    
    current_col = week_start_col
    for week_label, day1, day2, day3 in weeks_data:
        # Merge 3 cells for week header
        start_letter = get_column_letter(current_col)
        end_letter = get_column_letter(current_col + 2)
        ws.merge_cells(f'{start_letter}10:{end_letter}10')
        
        week_cell = ws[f'{start_letter}10']
        week_cell.value = week_label
        week_cell.fill = colors['light_gray']
        week_cell.font = Font(bold=True, size=10)
        week_cell.alignment = Alignment(horizontal='center', vertical='center')
        week_cell.border = thin_border
        
        # Add day headers in row 11
        for idx, day_text in enumerate([day1, day2, day3]):
            day_cell = ws.cell(row=11, column=current_col + idx)
            day_cell.value = day_text
            day_cell.fill = colors['light_gray']
            day_cell.font = Font(size=8)
            day_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            day_cell.border = thin_border
        
        current_col += 3
    
    # ===== MAIN HEADERS (Row 11) =====
    headers = [
        ('A11', 'No.', 5),
        ('B11', 'Activity Plan', 70),
        ('C11', 'Start Date', 11),
        ('D11', 'Target Date', 11),
        ('E11', 'Status', 10)
    ]
    
    for cell_ref, header_text, col_width in headers:
        cell = ws[cell_ref]
        cell.value = header_text
        cell.fill = colors['header_blue']
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Set column width
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = col_width
    
    # Set width for date columns
    for col_num in range(week_start_col, week_start_col + 12):
        ws.column_dimensions[get_column_letter(col_num)].width = 7
    
    # Set row heights
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 30
    
    # ===== PROJECT 1: Migne Realtime Plot Program =====
    activities_1 = [
        ('9', 'Migne Realtime Plot Program Modification', '20-Oct', '20-Oct', 'Done', 6, 1, 'green'),
        ('9.1', 'Changed method to automatically start scanning as soon as it starts', '20-Oct', '20-Oct', 'Done', 6, 1, 'green'),
        ('9.2', 'Integrating the GUI from the older scanning system interface', '21-Oct', '26-Oct', 'Done', 7, 3, 'green'),
        ('9.3', 'Testing w/ scanning system', '23-Oct', '25-Oct', 'Done', 8, 2, 'green'),
        ('9.4', 'For testing w/ scanning 10" & 12m clay', '26-Oct', '26-Oct', 'Done', 9, 1, 'green'),
        ('9.5', 'Implemented in the system - fullscreen + Assistant', '27-Oct', '27-Oct', 'Done', 10, 1, 'green'),
        ('9.6', 'Modifying new features', '28-Oct', '28-Oct', 'Done', 11, 1, 'green'),
        ('9.7', 'Modifying flash command program', '03-Nov', '08-Nov', 'Done', 13, 3, 'green'),
        ('9.7', 'Adjustment to the scanning function', '09-Nov', '09-Nov', 'Done', 15, 1, 'green'),
        ('9.9', 'Adding current date/time on function to the Realtime Plot & Overall Verification of the program', '30-Nov', '30-Nov', 'Ongoing', 16, 2, 'yellow'),
    ]
    
    start_row = 12
    for idx, (no, activity, start_date, target_date, status, gantt_col, gantt_span, gantt_color) in enumerate(activities_1):
        row = start_row + idx
        
        # Fill main columns
        ws.cell(row=row, column=1).value = no
        ws.cell(row=row, column=2).value = activity
        ws.cell(row=row, column=3).value = start_date
        ws.cell(row=row, column=4).value = target_date
        ws.cell(row=row, column=5).value = status
        
        # Format main columns
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(size=9)
            
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Status column coloring
            if col == 5:
                if status == 'Done':
                    cell.fill = colors['green']
                    cell.font = Font(size=9, color='FFFFFF', bold=True)
                elif status == 'Ongoing':
                    cell.fill = colors['orange']
                    cell.font = Font(size=9, bold=True)
        
        # Add Gantt bars
        for span_idx in range(gantt_span):
            gantt_cell = ws.cell(row=row, column=gantt_col + span_idx)
            gantt_cell.fill = colors[gantt_color]
            gantt_cell.border = thin_border
        
        # Fill empty date cells with borders
        for col in range(week_start_col, week_start_col + 12):
            cell = ws.cell(row=row, column=col)
            if cell.fill == colors['white'] or cell.fill.start_color.rgb == '00000000':
                cell.border = thin_border
        
        ws.row_dimensions[row].height = 18
    
    # ===== PROJECT 2: Plate Particle Cleaning =====
    project2_row = start_row + len(activities_1) + 2
    
    ws.cell(row=project2_row, column=1).value = '10'
    ws.cell(row=project2_row, column=2).value = 'Plate Particle Cleaning Evaluation'
    ws.cell(row=project2_row, column=2).font = Font(bold=True, size=10)
    ws.cell(row=project2_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    activities_2 = [
        ('10.1', 'Cleaning a piece of plate with cleaning methods from 1- P282', '10-Nov', '11-Nov', 'Done', 15, 2, 'green'),
        ('10.2', 'Cleaning another a piece recovered with experiment', '10-Nov', '11-Nov', 'Done', 15, 2, 'green'),
        ('10.3', 'Cleaning another a piece recovered with experiment', '10-Nov', '12-Nov', 'Done', 15, 2, 'green'),
    ]
    
    start_row_2 = project2_row + 1
    for idx, (no, activity, start_date, target_date, status, gantt_col, gantt_span, gantt_color) in enumerate(activities_2):
        row = start_row_2 + idx
        
        ws.cell(row=row, column=1).value = no
        ws.cell(row=row, column=2).value = activity
        ws.cell(row=row, column=3).value = start_date
        ws.cell(row=row, column=4).value = target_date
        ws.cell(row=row, column=5).value = status
        
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(size=9)
            
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col == 5:
                cell.fill = colors['green']
                cell.font = Font(size=9, color='FFFFFF', bold=True)
        
        for span_idx in range(gantt_span):
            gantt_cell = ws.cell(row=row, column=gantt_col + span_idx)
            gantt_cell.fill = colors[gantt_color]
            gantt_cell.border = thin_border
        
        for col in range(week_start_col, week_start_col + 12):
            cell = ws.cell(row=row, column=col)
            if cell.fill == colors['white'] or cell.fill.start_color.rgb == '00000000':
                cell.border = thin_border
        
        ws.row_dimensions[row].height = 18
    
    # Save file
    filename = 'Perfect_Gantt_Chart.xlsx'
    wb.save(filename)
    print(f"✓ Perfect Gantt chart created: {filename}")
    return filename

if __name__ == "__main__":
    create_gantt_chart()
