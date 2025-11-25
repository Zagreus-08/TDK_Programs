"""
Professional Gantt Chart Generator - Enhanced Edition v2
Features: Interactive menu, Add/Hide projects, Modern design
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
import os

# Data file to store projects
DATA_FILE = 'gantt_projects.json'

class GanttChartPro:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Activity Plan"
        
        # Modern Professional Color Palette
        self.colors = {
            'done': PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid'),
            'ongoing': PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid'),
            'plan': PatternFill(start_color='42A5F5', end_color='42A5F5', fill_type='solid'),
            'delayed': PatternFill(start_color='E53935', end_color='E53935', fill_type='solid'),
            'reschedule': PatternFill(start_color='AB47BC', end_color='AB47BC', fill_type='solid'),
            'on_hold': PatternFill(start_color='78909C', end_color='78909C', fill_type='solid'),
            
            'bar_done': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),
            'bar_ongoing': PatternFill(start_color='FFB74D', end_color='FFB74D', fill_type='solid'),
            'bar_plan': PatternFill(start_color='64B5F6', end_color='64B5F6', fill_type='solid'),
            'bar_delayed': PatternFill(start_color='EF5350', end_color='EF5350', fill_type='solid'),
            
            'header_primary': PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid'),
            'week_header': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
            'day_header': PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid'),
            'section_header': PatternFill(start_color='263238', end_color='263238', fill_type='solid'),
            'alt_row': PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid'),
            'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
            'today': PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid'),
            'weekend': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
        }
        
        self.thin_border = Border(
            left=Side(style='thin', color='BDBDBD'),
            right=Side(style='thin', color='BDBDBD'),
            top=Side(style='thin', color='BDBDBD'),
            bottom=Side(style='thin', color='BDBDBD')
        )
        
        self.current_row = 1
        self.date_col_start = 7
        self.dates = []

    def setup_page(self):
        self.ws.freeze_panes = 'G8'
        self.ws.sheet_view.showGridLines = False
        self.ws.column_dimensions['A'].width = 6
        self.ws.column_dimensions['B'].width = 8
        self.ws.column_dimensions['C'].width = 55
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 12
        self.ws.column_dimensions['F'].width = 12
        
    def create_title_section(self, title="ACTIVITY PLAN", subtitle="Weekly Progress Tracking Dashboard"):
        self.ws.merge_cells('A1:F1')
        title_cell = self.ws['A1']
        title_cell.value = title
        title_cell.font = Font(name='Calibri', size=24, bold=True, color='1565C0')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[1].height = 40
        
        self.ws.merge_cells('A2:F2')
        subtitle_cell = self.ws['A2']
        subtitle_cell.value = f"{subtitle} | Generated: {datetime.now().strftime('%B %d, %Y')}"
        subtitle_cell.font = Font(name='Calibri', size=11, italic=True, color='757575')
        subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
        self.ws.row_dimensions[2].height = 25
        self.current_row = 3
        
    def create_legend(self):
        row = self.current_row
        legend_items = [
            ('Done', 'done', '✓'),
            ('Ongoing', 'ongoing', '●'),
            ('Planned', 'plan', '○'),
            ('Delayed', 'delayed', '!'),
            ('Re-sch', 'reschedule', '↻'),
            ('On Hold', 'on_hold', '⏸'),
        ]
        
        col = 1
        for label, color_key, symbol in legend_items:
            cell = self.ws.cell(row=row, column=col)
            cell.value = symbol
            cell.fill = self.colors[color_key]
            cell.font = Font(size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            
            label_cell = self.ws.cell(row=row, column=col + 1)
            label_cell.value = label
            label_cell.font = Font(size=9, color='424242')
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            col += 2
        
        self.ws.row_dimensions[row].height = 22
        self.current_row = row + 2
        
    def generate_dates(self, start_date, num_weeks=8):
        self.dates = []
        current = start_date
        for _ in range(num_weeks * 7):
            self.dates.append(current)
            current += timedelta(days=1)
        return self.dates
    
    def create_date_headers(self, start_date, num_weeks=8):
        self.generate_dates(start_date, num_weeks)
        header_row = self.current_row
        day_row = header_row + 1
        
        col = self.date_col_start
        for week_num in range(num_weeks):
            week_start = start_date + timedelta(weeks=week_num)
            week_number = week_start.isocalendar()[1]
            
            start_col = get_column_letter(col)
            end_col = get_column_letter(col + 6)
            self.ws.merge_cells(f'{start_col}{header_row}:{end_col}{header_row}')
            
            week_cell = self.ws.cell(row=header_row, column=col)
            week_cell.value = f"WW{week_number}"
            week_cell.fill = self.colors['week_header']
            week_cell.font = Font(name='Calibri', size=11, bold=True, color='1565C0')
            week_cell.alignment = Alignment(horizontal='center', vertical='center')
            week_cell.border = self.thin_border
            
            for day_offset in range(7):
                day_date = week_start + timedelta(days=day_offset)
                day_cell = self.ws.cell(row=day_row, column=col + day_offset)
                day_cell.value = day_date.strftime('%d\n%a')
                
                if day_date.weekday() >= 5:
                    day_cell.fill = self.colors['weekend']
                    day_cell.font = Font(size=8, color='9E9E9E')
                else:
                    day_cell.fill = self.colors['day_header']
                    day_cell.font = Font(size=8, color='424242')
                
                if day_date.date() == datetime.now().date():
                    day_cell.fill = self.colors['today']
                    day_cell.font = Font(size=8, bold=True, color='F57F17')
                
                day_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                day_cell.border = self.thin_border
                self.ws.column_dimensions[get_column_letter(col + day_offset)].width = 5
            col += 7
        
        self.ws.row_dimensions[header_row].height = 22
        self.ws.row_dimensions[day_row].height = 30
        self.current_row = day_row + 1

    def create_main_headers(self):
        row = self.current_row
        headers = [('No.', 'A'), ('ID', 'B'), ('Activity / Task Description', 'C'),
                   ('Start', 'D'), ('End', 'E'), ('Status', 'F')]
        
        for header_text, col_letter in headers:
            cell = self.ws[f'{col_letter}{row}']
            cell.value = header_text
            cell.fill = self.colors['header_primary']
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        self.ws.row_dimensions[row].height = 28
        self.current_row = row + 1
        
    def add_section_header(self, section_num, section_title):
        row = self.current_row
        
        num_cell = self.ws.cell(row=row, column=1)
        num_cell.value = section_num
        num_cell.fill = self.colors['section_header']
        num_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.border = self.thin_border
        
        self.ws.merge_cells(f'B{row}:F{row}')
        title_cell = self.ws.cell(row=row, column=2)
        title_cell.value = f"  {section_title}"
        title_cell.fill = self.colors['section_header']
        title_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        title_cell.border = self.thin_border
        
        for col in range(self.date_col_start, self.date_col_start + len(self.dates)):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = self.colors['section_header']
            cell.border = self.thin_border
        
        self.ws.row_dimensions[row].height = 26
        self.current_row = row + 1
        
    def add_activity(self, row_num, task_id, description, start_date, end_date, status):
        row = self.current_row
        is_alt_row = (row % 2 == 0)
        
        status_config = {
            'Done': ('done', 'bar_done', '✓', 'FFFFFF'),
            'Ongoing': ('ongoing', 'bar_ongoing', '●', 'FFFFFF'),
            'Plan': ('plan', 'bar_plan', '○', 'FFFFFF'),
            'Planned': ('plan', 'bar_plan', '○', 'FFFFFF'),
            'Delayed': ('delayed', 'bar_delayed', '!', 'FFFFFF'),
            'Re-schedule': ('reschedule', 'bar_plan', '↻', 'FFFFFF'),
            'On Hold': ('on_hold', 'bar_plan', '⏸', 'FFFFFF'),
        }
        
        status_fill, bar_color, status_symbol, text_color = status_config.get(
            status, ('plan', 'bar_plan', '○', 'FFFFFF'))
        
        row_fill = self.colors['alt_row'] if is_alt_row else self.colors['white']
        
        # Row number
        cell_num = self.ws.cell(row=row, column=1)
        cell_num.value = row_num
        cell_num.fill = row_fill
        cell_num.font = Font(size=9, color='757575')
        cell_num.alignment = Alignment(horizontal='center', vertical='center')
        cell_num.border = self.thin_border
        
        # Task ID
        cell_id = self.ws.cell(row=row, column=2)
        cell_id.value = task_id
        cell_id.fill = row_fill
        cell_id.font = Font(size=9, bold=True, color='424242')
        cell_id.alignment = Alignment(horizontal='center', vertical='center')
        cell_id.border = self.thin_border
        
        # Description
        cell_desc = self.ws.cell(row=row, column=3)
        cell_desc.value = description
        cell_desc.fill = row_fill
        cell_desc.font = Font(size=9, color='212121')
        cell_desc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell_desc.border = self.thin_border
        
        # Start date
        cell_start = self.ws.cell(row=row, column=4)
        cell_start.value = start_date.strftime('%d-%b') if isinstance(start_date, datetime) else start_date
        cell_start.fill = row_fill
        cell_start.font = Font(size=9, color='616161')
        cell_start.alignment = Alignment(horizontal='center', vertical='center')
        cell_start.border = self.thin_border
        
        # End date
        cell_end = self.ws.cell(row=row, column=5)
        cell_end.value = end_date.strftime('%d-%b') if isinstance(end_date, datetime) else end_date
        cell_end.fill = row_fill
        cell_end.font = Font(size=9, color='616161')
        cell_end.alignment = Alignment(horizontal='center', vertical='center')
        cell_end.border = self.thin_border
        
        # Status
        cell_status = self.ws.cell(row=row, column=6)
        cell_status.value = f"{status_symbol} {status}"
        cell_status.fill = self.colors[status_fill]
        cell_status.font = Font(size=9, bold=True, color=text_color)
        cell_status.alignment = Alignment(horizontal='center', vertical='center')
        cell_status.border = self.thin_border
        
        self._draw_gantt_bar(row, start_date, end_date, bar_color, row_fill)
        self.ws.row_dimensions[row].height = 22
        self.current_row = row + 1

    def _draw_gantt_bar(self, row, start_date, end_date, bar_color, row_fill):
        if not self.dates:
            return
            
        if isinstance(start_date, str):
            try:
                start_date = datetime.strptime(f"{start_date}-2025", '%d-%b-%Y')
            except:
                return
        if isinstance(end_date, str):
            try:
                end_date = datetime.strptime(f"{end_date}-2025", '%d-%b-%Y')
            except:
                return
        
        for col_idx, date in enumerate(self.dates):
            cell = self.ws.cell(row=row, column=self.date_col_start + col_idx)
            if date.weekday() >= 5:
                cell.fill = self.colors['weekend']
            else:
                cell.fill = row_fill
            if date.date() == datetime.now().date():
                cell.fill = self.colors['today']
            cell.border = self.thin_border
        
        for col_idx, date in enumerate(self.dates):
            if start_date.date() <= date.date() <= end_date.date():
                cell = self.ws.cell(row=row, column=self.date_col_start + col_idx)
                cell.fill = self.colors[bar_color]
                cell.border = self.thin_border
                
    def add_empty_row(self):
        self.ws.row_dimensions[self.current_row].height = 8
        self.current_row += 1
        
    def save(self, filename='Professional_Gantt_Chart.xlsx'):
        self.wb.save(filename)
        print(f"✅ Gantt chart saved: {filename}")
        return filename


# ============== PROJECT DATA MANAGEMENT ==============

def load_projects():
    """Load projects from JSON file"""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"projects": {}, "settings": {"start_date": "2025-11-10", "num_weeks": 6}}

def save_projects(data):
    """Save projects to JSON file"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("💾 Projects saved!")

def add_project(data):
    """Add a new project"""
    print("\n" + "="*50)
    print("➕ ADD NEW PROJECT")
    print("="*50)
    
    proj_num = input("Enter project number (e.g., 9, 10, 11): ").strip()
    proj_name = input("Enter project name: ").strip()
    
    if not proj_num or not proj_name:
        print("❌ Project number and name are required!")
        return
    
    data["projects"][proj_num] = {
        "name": proj_name,
        "visible": True,
        "tasks": []
    }
    
    print(f"✅ Project {proj_num}: '{proj_name}' created!")
    
    # Ask to add tasks
    while True:
        add_task = input("\nAdd a task to this project? (y/n): ").strip().lower()
        if add_task != 'y':
            break
        add_task_to_project(data, proj_num)
    
    save_projects(data)

def add_task_to_project(data, proj_num=None):
    """Add a task to a project"""
    if proj_num is None:
        print("\nAvailable projects:")
        for num, proj in data["projects"].items():
            status = "👁️" if proj["visible"] else "🙈"
            print(f"  {status} [{num}] {proj['name']} ({len(proj['tasks'])} tasks)")
        proj_num = input("Enter project number: ").strip()
    
    if proj_num not in data["projects"]:
        print("❌ Project not found!")
        return
    
    proj = data["projects"][proj_num]
    task_count = len(proj["tasks"]) + 1
    
    print(f"\n📝 Adding task to Project {proj_num}: {proj['name']}")
    
    task_id = input(f"Task ID (default: {proj_num}.{task_count}): ").strip()
    if not task_id:
        task_id = f"{proj_num}.{task_count}"
    
    description = input("Task description: ").strip()
    start_date = input("Start date (DD-Mon, e.g., 25-Nov): ").strip()
    end_date = input("End date (DD-Mon, e.g., 27-Nov): ").strip()
    
    print("\nStatus options: Done, Ongoing, Plan, Delayed, Re-schedule, On Hold")
    status = input("Status (default: Plan): ").strip()
    if not status:
        status = "Plan"
    
    proj["tasks"].append({
        "task_id": task_id,
        "description": description,
        "start_date": start_date,
        "end_date": end_date,
        "status": status
    })
    
    print(f"✅ Task '{task_id}' added!")

def toggle_project_visibility(data):
    """Show/Hide projects"""
    print("\n" + "="*50)
    print("👁️ SHOW/HIDE PROJECTS")
    print("="*50)
    
    if not data["projects"]:
        print("No projects yet!")
        return
    
    print("\nCurrent projects:")
    for num, proj in sorted(data["projects"].items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
        status = "✅ Visible" if proj["visible"] else "❌ Hidden"
        print(f"  [{num}] {proj['name']} - {status} ({len(proj['tasks'])} tasks)")
    
    proj_num = input("\nEnter project number to toggle visibility: ").strip()
    
    if proj_num in data["projects"]:
        data["projects"][proj_num]["visible"] = not data["projects"][proj_num]["visible"]
        new_status = "visible" if data["projects"][proj_num]["visible"] else "hidden"
        print(f"✅ Project {proj_num} is now {new_status}")
        save_projects(data)
    else:
        print("❌ Project not found!")

def list_projects(data):
    """List all projects and their tasks"""
    print("\n" + "="*50)
    print("📋 ALL PROJECTS")
    print("="*50)
    
    if not data["projects"]:
        print("No projects yet! Use option 1 to add a project.")
        return
    
    for num, proj in sorted(data["projects"].items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
        status = "👁️ VISIBLE" if proj["visible"] else "🙈 HIDDEN"
        print(f"\n[{num}] {proj['name']} - {status}")
        print("-" * 40)
        
        if proj["tasks"]:
            for i, task in enumerate(proj["tasks"], 1):
                print(f"  {i}. [{task['task_id']}] {task['description'][:40]}...")
                print(f"     📅 {task['start_date']} → {task['end_date']} | {task['status']}")
        else:
            print("  (No tasks yet)")

def edit_settings(data):
    """Edit chart settings"""
    print("\n" + "="*50)
    print("⚙️ CHART SETTINGS")
    print("="*50)
    
    print(f"\nCurrent settings:")
    print(f"  Start date: {data['settings']['start_date']}")
    print(f"  Number of weeks: {data['settings']['num_weeks']}")
    
    new_start = input("\nNew start date (YYYY-MM-DD) or press Enter to keep: ").strip()
    if new_start:
        data['settings']['start_date'] = new_start
    
    new_weeks = input("Number of weeks to display (4-12) or press Enter to keep: ").strip()
    if new_weeks and new_weeks.isdigit():
        data['settings']['num_weeks'] = int(new_weeks)
    
    save_projects(data)
    print("✅ Settings updated!")

def delete_project(data):
    """Delete a project"""
    print("\n" + "="*50)
    print("🗑️ DELETE PROJECT")
    print("="*50)
    
    list_projects(data)
    
    proj_num = input("\nEnter project number to delete: ").strip()
    
    if proj_num in data["projects"]:
        confirm = input(f"Are you sure you want to delete project {proj_num}? (yes/no): ").strip().lower()
        if confirm == 'yes':
            del data["projects"][proj_num]
            save_projects(data)
            print(f"✅ Project {proj_num} deleted!")
        else:
            print("Cancelled.")
    else:
        print("❌ Project not found!")

def generate_chart(data):
    """Generate the Excel Gantt chart"""
    print("\n" + "="*50)
    print("📊 GENERATING GANTT CHART")
    print("="*50)
    
    # Filter visible projects
    visible_projects = {k: v for k, v in data["projects"].items() if v["visible"]}
    
    if not visible_projects:
        print("❌ No visible projects! Show at least one project first.")
        return
    
    gantt = GanttChartPro()
    gantt.setup_page()
    gantt.create_title_section("ACTIVITY PLAN", "Weekly Progress Tracking Dashboard")
    gantt.create_legend()
    
    # Parse start date
    start_date = datetime.strptime(data['settings']['start_date'], '%Y-%m-%d')
    num_weeks = data['settings']['num_weeks']
    
    gantt.create_date_headers(start_date, num_weeks)
    gantt.create_main_headers()
    
    # Add each visible project
    for proj_num, proj in sorted(visible_projects.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
        gantt.add_section_header(proj_num, proj["name"])
        
        for i, task in enumerate(proj["tasks"], 1):
            # Parse dates
            try:
                task_start = datetime.strptime(f"{task['start_date']}-2025", '%d-%b-%Y')
                task_end = datetime.strptime(f"{task['end_date']}-2025", '%d-%b-%Y')
            except:
                task_start = task['start_date']
                task_end = task['end_date']
            
            gantt.add_activity(i, task['task_id'], task['description'], 
                             task_start, task_end, task['status'])
        
        gantt.add_empty_row()
    
    filename = 'Professional_Gantt_Chart.xlsx'
    gantt.save(filename)
    print(f"\n🎉 Chart generated with {len(visible_projects)} project(s)!")
    print(f"📁 Open: {filename}")

def add_sample_data(data):
    """Add sample projects for demonstration"""
    data["projects"] = {
        "9": {
            "name": "Migne Realtime Plot Program Modification",
            "visible": True,
            "tasks": [
                {"task_id": "9.1", "description": "Changed function to automatically start instead of waiting 1st on data", "start_date": "10-Nov", "end_date": "10-Nov", "status": "Done"},
                {"task_id": "9.2", "description": "Integrating the GUI from the older Scanning System Interface", "start_date": "10-Nov", "end_date": "12-Nov", "status": "Done"},
                {"task_id": "9.3", "description": "Testing on scanning system", "start_date": "13-Nov", "end_date": "14-Nov", "status": "Done"},
                {"task_id": "9.4", "description": "For review with S. Estrada", "start_date": "14-Nov", "end_date": "14-Nov", "status": "Done"},
                {"task_id": "9.5", "description": "Implemented in the system - Fullscreen + Autostart", "start_date": "17-Nov", "end_date": "17-Nov", "status": "Done"},
                {"task_id": "9.6", "description": "Modifying new features", "start_date": "18-Nov", "end_date": "19-Nov", "status": "Done"},
                {"task_id": "9.7", "description": "Modifying main command program", "start_date": "20-Nov", "end_date": "21-Nov", "status": "Done"},
                {"task_id": "9.8", "description": "Adjustment to the scanning functions", "start_date": "24-Nov", "end_date": "25-Nov", "status": "Ongoing"},
                {"task_id": "9.9", "description": "Scanning Plate at FR=0.13, 0.15, & 0.18", "start_date": "26-Nov", "end_date": "26-Nov", "status": "Plan"},
                {"task_id": "9.10", "description": "Adding Z-range adjustment function to the Realtime Plot & Overall Modifications", "start_date": "27-Nov", "end_date": "01-Dec", "status": "Plan"},
            ]
        },
        "10": {
            "name": "Plate Particle Cleaning Evaluation",
            "visible": True,
            "tasks": [
                {"task_id": "10.1", "description": "Cleaning 3 pieces of plate with cleaning methods from 3 Pdoc", "start_date": "10-Nov", "end_date": "11-Nov", "status": "Done"},
                {"task_id": "10.2", "description": "Cleaning another 3 pieces recovered by Kasahara-san", "start_date": "11-Nov", "end_date": "12-Nov", "status": "Done"},
                {"task_id": "10.3", "description": "Give the 3 pieces result to Kasahara-san & Kamei-san", "start_date": "12-Nov", "end_date": "12-Nov", "status": "Done"},
                {"task_id": "10.4", "description": "For review with Kasajima san & Kami san", "start_date": "13-Nov", "end_date": "13-Nov", "status": "Done"},
                {"task_id": "10.5", "description": "Implemented in the system - Fullscreen + Autostart", "start_date": "17-Nov", "end_date": "18-Nov", "status": "Done"},
                {"task_id": "10.6", "description": "Modifying new features", "start_date": "19-Nov", "end_date": "20-Nov", "status": "Ongoing"},
                {"task_id": "10.7", "description": "Modifying main command program", "start_date": "24-Nov", "end_date": "26-Nov", "status": "Plan"},
                {"task_id": "10.8", "description": "Adjustment to the scanning functions", "start_date": "27-Nov", "end_date": "28-Nov", "status": "Plan"},
            ]
        },
        "11": {
            "name": "Auto Coil Winding Improvement",
            "visible": True,
            "tasks": [
                {"task_id": "11.1", "description": "Initial assessment and planning", "start_date": "17-Nov", "end_date": "18-Nov", "status": "Done"},
                {"task_id": "11.2", "description": "Design modifications review", "start_date": "19-Nov", "end_date": "21-Nov", "status": "Ongoing"},
                {"task_id": "11.3", "description": "Implementation of changes", "start_date": "24-Nov", "end_date": "28-Nov", "status": "Plan"},
                {"task_id": "11.4", "description": "Testing and validation", "start_date": "01-Dec", "end_date": "05-Dec", "status": "Plan"},
            ]
        }
    }
    save_projects(data)
    print("✅ Sample data loaded!")


def main_menu():
    """Main interactive menu"""
    print("\n" + "="*60)
    print("   📊 PROFESSIONAL GANTT CHART GENERATOR v2.0")
    print("="*60)
    
    data = load_projects()
    
    while True:
        print("\n┌─────────────────────────────────────┐")
        print("│           MAIN MENU                 │")
        print("├─────────────────────────────────────┤")
        print("│  1. ➕ Add New Project              │")
        print("│  2. 📝 Add Task to Project          │")
        print("│  3. 👁️  Show/Hide Projects          │")
        print("│  4. 📋 List All Projects            │")
        print("│  5. 🗑️  Delete Project              │")
        print("│  6. ⚙️  Chart Settings              │")
        print("│  7. 📊 Generate Gantt Chart         │")
        print("│  8. 📥 Load Sample Data             │")
        print("│  0. 🚪 Exit                         │")
        print("└─────────────────────────────────────┘")
        
        choice = input("\nEnter your choice (0-8): ").strip()
        
        if choice == '1':
            add_project(data)
        elif choice == '2':
            add_task_to_project(data)
            save_projects(data)
        elif choice == '3':
            toggle_project_visibility(data)
        elif choice == '4':
            list_projects(data)
        elif choice == '5':
            delete_project(data)
        elif choice == '6':
            edit_settings(data)
        elif choice == '7':
            generate_chart(data)
        elif choice == '8':
            add_sample_data(data)
        elif choice == '0':
            print("\n👋 Goodbye!")
            break
        else:
            print("❌ Invalid choice. Please try again.")


if __name__ == "__main__":
    main_menu()
